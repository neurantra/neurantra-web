#!/usr/bin/env python3
"""Build the infrastructure tracker from live sources, not from guesses.

    uv run --with openpyxl python scripts/build-infra-tracker.py [--out PATH]

The spreadsheet this replaces was produced by scanning ~/dev and inferring each project's
hosting from the files it found. That is the wrong instrument for the question. A render.yaml
in a tree is evidence that something was once deployed to Render; it is not evidence of where
traffic goes today. It made the scanner report that indiamasala deploys to Render (that config
was ported to GitHub Actions and deleted) and that surgerycare-web is on AWS (its backend is;
the site is Cloudflare Pages) — both stated with the same confidence as the rows that happened
to be right.

So every cell here carries its provenance:

    live      queried from the GitHub API at generation time. Cannot go stale.
    repo      read from the working tree on this machine. Stale only if the clone is.
    declared  a human wrote it in infra-declared.json, because nothing can confirm it.

Correcting a declared fact means editing that JSON and re-running. Correcting a live one is
impossible, which is the point. Hand-editing the .xlsx is the one thing not to do: the next
run overwrites it, which is exactly how the last tracker lost its corrections.

Local clones are optional. Without them the repo-derived columns read "no local clone" and
everything else still generates.
"""

from __future__ import annotations

import argparse
import json
import pathlib
import re
import subprocess
import sys
from datetime import datetime

HERE = pathlib.Path(__file__).resolve().parent
DECLARED = HERE / "infra-declared.json"
DEV = pathlib.Path.home() / "dev"

HDR_BG = "1F4B7A"


def gh(path: str):
    """One GitHub API call. Returns None rather than raising: a tracker that refuses to
    generate because one endpoint is unavailable is less useful than one with a gap in it."""
    r = subprocess.run(["gh", "api", path], capture_output=True, text=True)
    if r.returncode:
        return None
    try:
        return json.loads(r.stdout)
    except json.JSONDecodeError:
        return None


def clone_for(repo: str) -> pathlib.Path | None:
    """Find the local clone by remote, not by folder name — they have disagreed before."""
    for d in sorted(DEV.iterdir()) if DEV.is_dir() else []:
        if not (d / ".git").is_dir():
            continue
        r = subprocess.run(["git", "-C", str(d), "remote", "get-url", "origin"],
                           capture_output=True, text=True)
        if r.returncode == 0 and r.stdout.strip().rstrip("/").split("/")[-1].removesuffix(".git") == repo:
            return d
    # surgerycare-app lives inside another repo's tree, so the top-level sweep misses it
    for g in DEV.glob("*/*/*/*/.git"):
        d = g.parent
        r = subprocess.run(["git", "-C", str(d), "remote", "get-url", "origin"],
                           capture_output=True, text=True)
        if r.returncode == 0 and r.stdout.strip().rstrip("/").split("/")[-1].removesuffix(".git") == repo:
            return d
    return None


def current_branch(d: pathlib.Path) -> str:
    r = subprocess.run(["git", "-C", str(d), "branch", "--show-current"],
                       capture_output=True, text=True)
    return r.stdout.strip() or "?"


def app_version(path: pathlib.Path) -> tuple[str, str, str]:
    """(framework, version, build) from whichever file this app keeps it in."""
    pub = path / "pubspec.yaml"
    if pub.is_file():
        m = re.search(r"^version:\s*(\S+)", pub.read_text(), re.M)
        if m and "+" in m.group(1):
            name, build = m.group(1).split("+", 1)
            return "Flutter", name, build
        return "Flutter", m.group(1) if m else "?", "?"
    aj = path / "app.json"
    if aj.is_file():
        try:
            e = json.loads(aj.read_text())["expo"]
        except Exception:
            return "Expo", "?", "?"
        sdk = ""
        pkg = path / "package.json"
        if pkg.is_file():
            try:
                sdk = json.loads(pkg.read_text()).get("dependencies", {}).get("expo", "")
            except Exception:
                pass
        return f"Expo {sdk}".strip(), str(e.get("version", "?")), str(e.get("ios", {}).get("buildNumber", "?"))
    return "?", "?", "?"


def native_layout(path: pathlib.Path) -> str:
    """Whether ios/ and android/ are committed or regenerated — it decides what a version
    bump has to touch, and it differs per platform in at least one app."""
    def ignored(sub: str) -> bool:
        return subprocess.run(["git", "-C", str(path), "check-ignore", "-q", sub]).returncode == 0
    ios = "generated" if ignored("ios") else "committed"
    and_ = "generated" if ignored("android") else "committed"
    return f"ios {ios}, android {and_}"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=str(DEV / "outputs" / "infra-tracker.xlsx"))
    args = ap.parse_args()

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl is needed:  uv run --with openpyxl python scripts/build-infra-tracker.py",
              file=sys.stderr)
        return 1

    cfg = json.loads(DECLARED.read_text())
    org = cfg["org"]

    repos = gh(f"/orgs/{org}/repos?per_page=100")
    if repos is None:
        print(f"could not list repos for {org} — is `gh` authenticated?", file=sys.stderr)
        return 1
    repos.sort(key=lambda r: r["name"])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def sheet(name, headers, rows, widths):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for c in ws[1]:
            c.fill = PatternFill("solid", fgColor=HDR_BG)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.alignment = Alignment(vertical="center", wrap_text=True)
        for r in rows:
            ws.append(r)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = Alignment(vertical="top", wrap_text=True)
        return ws

    # ── Repos ────────────────────────────────────────────────────────────────────
    rows, mobile_rows, secret_rows = [], [], []
    for r in repos:
        name = r["name"]
        d = cfg["repos"].get(name, {})
        clone = clone_for(name)
        envs = gh(f"/repos/{org}/{name}/environments") or {}
        wf = (gh(f"/repos/{org}/{name}/actions/workflows") or {}).get("total_count", 0)
        prot = [b["name"] for b in (gh(f"/repos/{org}/{name}/branches?per_page=50") or [])
                if b.get("protected")]
        rows.append([
            name,
            d.get("purpose") or "",
            r["default_branch"],
            current_branch(clone) if clone else "no local clone",
            r.get("language") or "",
            d.get("hosting") or "",
            d.get("trigger") or "",
            d.get("root_dir") or "",
            d.get("url") or "",
            wf,
            (envs.get("total_count") or 0),
            ", ".join(prot) or "none",
            r["pushed_at"][:10],
            f"{round(r['size'] / 1024)} MB",
            d.get("extra") or "",
        ])

        # secrets: names and location only. A tracker must never carry a value.
        for e in (envs.get("environments") or []):
            for s in ((gh(f"/repos/{org}/{name}/environments/{e['name']}/secrets") or {}).get("secrets") or []):
                secret_rows.append([name, f"environment: {e['name']}", s["name"], "GitHub environment secret", s["created_at"][:10]])
        for s in ((gh(f"/repos/{org}/{name}/actions/secrets") or {}).get("secrets") or []):
            secret_rows.append([name, "repository", s["name"], "GitHub repository secret", s["created_at"][:10]])
        for v in ((gh(f"/repos/{org}/{name}/actions/variables") or {}).get("variables") or []):
            secret_rows.append([name, "repository", v["name"], "GitHub variable (not secret — deliberately readable)", v["created_at"][:10]])

    sheet("Repos",
          ["Repo", "Purpose [declared]", "Prod branch [live]", "Local branch [repo]",
           "Language [live]", "Hosting [declared]", "Trigger [declared]", "Root dir [declared]",
           "Live URL [declared]", "Workflows [live]", "Envs [live]", "Protected [live]",
           "Last push [live]", "Size [live]", "Notes [declared]"],
          rows,
          [24, 40, 14, 15, 12, 30, 15, 24, 26, 11, 8, 12, 12, 9, 70])

    # ── Mobile ───────────────────────────────────────────────────────────────────
    for m in cfg["mobile_apps"]:
        clone = clone_for(m["repo"])
        path = (clone / m["path"]).resolve() if clone else None
        if path and path.is_dir():
            fw, ver, build = app_version(path)
            layout = native_layout(path)
            rp = "yes" if (path / "scripts" / "release-prep.sh").is_file() else "MISSING"
            ba = "yes" if (path / "scripts" / "build-android.sh").is_file() else "MISSING"
        else:
            fw = ver = build = layout = rp = ba = "no local clone"
        mobile_rows.append([m["app"], m["repo"], m["path"], fw, ver, build, layout, rp, ba, m["signing"]])

    sheet("Mobile",
          ["App", "Repo [declared]", "Path [declared]", "Framework [repo]", "Version [repo]",
           "iOS build [repo]", "Native layout [repo]", "release-prep.sh [repo]",
           "build-android.sh [repo]", "Android signing [declared]"],
          mobile_rows,
          [16, 24, 14, 22, 10, 11, 30, 20, 21, 62])

    # ── Services, Secrets, Attention ────────────────────────────────────────────
    sheet("Services",
          ["Service", "Used by", "Billed to", "Limit / budget", "Note"],
          [[s["service"], s["used_by"], s["billed_to"], s["limit"], s.get("note") or ""]
           for s in cfg["services"]],
          [20, 54, 20, 44, 70])

    secret_rows.sort(key=lambda r: (r[0], r[1], r[2]))
    sheet("Secrets",
          ["Repo [live]", "Scope [live]", "Name [live]", "Kind [live]", "Created [live]"],
          secret_rows,
          [24, 26, 34, 46, 12])

    sheet("Attention",
          ["Item", "Where", "Why it matters", "Action"],
          [[a["item"], a["repos"], a["why"], a["action"]] for a in cfg["attention"]],
          [42, 34, 78, 52])

    # ── How this is made ────────────────────────────────────────────────────────
    ws = wb.create_sheet("How this is made", 0)
    for a, b in [
        ("Neurantra infrastructure tracker", ""),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Generated by", "neurantra-web/scripts/build-infra-tracker.py"),
        ("Declared facts", "neurantra-web/scripts/infra-declared.json"),
        ("Regenerate", "uv run --with openpyxl python scripts/build-infra-tracker.py"),
        ("", ""),
        ("Do not hand-edit this file", "The next run overwrites it. Edit infra-declared.json and regenerate."),
        ("", ""),
        ("Column tags", ""),
        ("  [live]", "Queried from the GitHub API when this was generated. Cannot be stale."),
        ("  [repo]", "Read from the working tree on this machine. Stale only if the clone is."),
        ("  [declared]", "A human wrote it in infra-declared.json. Nothing can confirm it automatically."),
        ("", ""),
        ("Why the split", "The previous tracker inferred hosting from files on disk and got two repos"),
        ("", "wrong the same way: a render.yaml made it say indiamasala deploys to Render, and"),
        ("", "a main.tf made it say surgerycare-web is on AWS. A file is evidence something was"),
        ("", "once true, not evidence of where traffic goes now. Those facts are declared here"),
        ("", "so a wrong one is someone's mistake rather than a heuristic's."),
        ("", ""),
        ("Shipping runbook", "neurantra-web/docs/SHIPPING.md — how to deploy. This tracks what exists."),
        ("Repos", len(repos)),
        ("Mobile apps", len(cfg["mobile_apps"])),
        ("Secrets catalogued", len(secret_rows)),
    ]:
        ws.append([a, b])
    ws["A1"].font = Font(bold=True, size=13)
    ws["A7"].font = Font(bold=True, color="94382C")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 92

    out = pathlib.Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")
    print(f"  {len(repos)} repos, {len(cfg['mobile_apps'])} apps, {len(secret_rows)} secrets, "
          f"{len(cfg['services'])} services, {len(cfg['attention'])} attention items")
    missing = [m[0] for m in mobile_rows if "MISSING" in (m[7], m[8])]
    if missing:
        print(f"  apps missing a script: {', '.join(missing)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
